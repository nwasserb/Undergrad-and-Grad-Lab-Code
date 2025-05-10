import pyvisa as visa
import time
import math
from time import sleep, time as current_time
import serial
import threading
import sys
import datetime
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt

# Initialize Excel workbook and worksheet for real-time data logging
excel_filename = "noise_log.xlsx"
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Noise Data"
    # Write header row with actual values (dBm, mW) and the date
    ws.append(["Timestamp", "Date", "Port1 (dBm)", "Port2 (dBm)", "Port3 (dBm)", "Port4 (dBm)",
               "Port1 (mW)", "Port2 (mW)", "Port3 (mW)", "Port4 (mW)", "Total mW"])
    wb.save(excel_filename)
except Exception as e:
    print("Error initializing Excel file:", e)
    sys.exit(1)

def update_excel(timestamp, port1_dBm, port2_dBm, port3_dBm, port4_dBm,
                 port1_mW, port2_mW, port3_mW, port4_mW, total_mW):
    """Append a row of data to the Excel file, including the date."""
    try:
        # Extract the date portion from the timestamp (YYYY-MM-DD)
        date_str = datetime.datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        ws.append([timestamp, date_str, port1_dBm, port2_dBm, port3_dBm, port4_dBm,
                   port1_mW, port2_mW, port3_mW, port4_mW, total_mW])
        wb.save(excel_filename)
    except Exception as e:
        print("Error updating Excel file:", e)

def continuous_monitoring(session, run_duration):
    """
    Continuously monitor the detector, log actual data in real time to Excel, and print measurements.
    run_duration: Total duration for logging (in seconds)
    **CHANGE HERE: Set run_duration to your desired duration, e.g. 2 days = 172800 seconds**
    """
    start_time = current_time()
    while current_time() - start_time < run_duration:
        try:
            result = session.query('READ? 0')
            ports = result.split(',')
            # Parse actual dBm readings for each port.
            port1_dBm = float(ports[0].split(':')[-1])
            port2_dBm = float(ports[1])
            port3_dBm = float(ports[2])
            port4_dBm = float(ports[3])
            
            # Convert dBm readings to mW
            port1_mW = pow(10, port1_dBm / 10)
            port2_mW = pow(10, port2_dBm / 10)
            port3_mW = pow(10, port3_dBm / 10)
            port4_mW = pow(10, port4_dBm / 10)
            total_mW = port1_mW + port2_mW + port3_mW + port4_mW
            
            # Print the actual values (dBm)
            sys.stdout.write("\r\033[K")
            sys.stdout.write(f"Port1: {port1_dBm:6.2f} dBm | Port2: {port2_dBm:6.2f} dBm | "
                             f"Port3: {port3_dBm:6.2f} dBm | Port4: {port4_dBm:6.2f} dBm")
            sys.stdout.flush()
            
            # Log data to Excel file with timestamp
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            update_excel(timestamp, port1_dBm, port2_dBm, port3_dBm, port4_dBm,
                         port1_mW, port2_mW, port3_mW, port4_mW, total_mW)
            
            sleep(0.1)  # adjust as needed for sampling frequency
        except visa.VisaIOError as e:
            print(f'\nError: {e}')
            break
        except Exception as ex:
            print("Unexpected error during monitoring:", ex)
            break

def plot_data():
    """Load the Excel data and plot the actual dBm measurements over time."""
    try:
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        timestamps = []
        port1_vals = []
        port2_vals = []
        port3_vals = []
        port4_vals = []
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            ts, day, p1_dBm, p2_dBm, p3_dBm, p4_dBm, _, _, _, _, _ = row
            timestamps.append(datetime.datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"))
            port1_vals.append(p1_dBm)
            port2_vals.append(p2_dBm)
            port3_vals.append(p3_dBm)
            port4_vals.append(p4_dBm)
        
        plt.figure(figsize=(12, 6))
        plt.plot(timestamps, port1_vals, label='Port1 (dBm)')
        plt.plot(timestamps, port2_vals, label='Port2 (dBm)')
        plt.plot(timestamps, port3_vals, label='Port3 (dBm)')
        plt.plot(timestamps, port4_vals, label='Port4 (dBm)')
        plt.xlabel("Time")
        plt.ylabel("Measured Value (dBm)")
        plt.title("System Noise Measurements Over Time")
        plt.legend()
        plt.tight_layout()
        plt.show()
    except Exception as e:
        print("Error plotting data:", e)

def main():
    # ---------------------------
    # Open VISA session
    # ---------------------------
    try:
        resourceManager = visa.ResourceManager()
        dev = 'TCPIP0::10.0.3.17::5000::SOCKET'
        session = resourceManager.open_resource(dev)
        session.read_termination = '\n'
        session.write_termination = '\n'
        print("VISA Session Opened. Resources:", resourceManager.list_resources())
    except visa.VisaIOError as e:
        print(f"Error opening VISA session: {e}")
        return

    # If you don't need serial communication for basis setup, you can remove or comment this section.
    try:
        ser = serial.Serial(
            port='COM33',
            baudrate=19200,
            bytesize=serial.EIGHTBITS,
            parity=serial.PARITY_NONE,
            stopbits=serial.STOPBITS_TWO,
            timeout=1
        )
    except Exception as e:
        print("Error opening serial port:", e)
        session.close()
        return

    # If no basis initialization is required, skip these sections.
    
    # ---------------------------
    # Continuous monitoring & data logging
    # ---------------------------
    # **CHANGE HERE: Set run_duration to desired time in seconds (2 days = 172800 seconds)**
    # run_duration = 172800  # 2 days in seconds
    run_duration = 86400
    print(f"Starting continuous monitoring for {run_duration} seconds...")
    continuous_monitoring(session, run_duration)
    
    # Close the VISA session once done
    session.close()
    
    # ---------------------------
    # Plot the logged data
    # ---------------------------
    plot_data()

if __name__ == "__main__":
    main()
